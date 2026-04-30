import sys
from typing import Generator
import asyncio
from pathlib import Path
import openpyxl
from deep_translator import GoogleTranslator


PROJECT_ROOT = str(Path(__file__).resolve().parents[1])
if PROJECT_ROOT not in sys.path:
	sys.path.insert(0, PROJECT_ROOT)

async def translator(name: str) -> str:
	if name is None:
		return None
	translated = await asyncio.to_thread(
		GoogleTranslator(source="ru", target="en").translate, name.strip()
	)
	return translated.lower().replace(" ", "-").replace("_", "-")


def open_xlsx_file(file_path: str) -> Generator:
	wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
	ws = wb.active
	try:
		for row in ws.iter_rows(values_only=True):
			yield list(row)
	finally:
		wb.close()


async def main(file_path: str) -> None:
	trans = openpyxl.Workbook()
	trans_s = trans.active
	l_row, l_row_t = [None] * 100, [None] * 100
	p = 0
	for row in open_xlsx_file(file_path):
		p += 1
		row_t = []
		for i in range(len(row)):
			if row[i] == l_row[i]:
				row_t.append(l_row_t[i])
			else:
				row_t.append(await translator(row[i]))
		trans_s.append(row_t)
		l_row ,l_row_t = row, row_t
		print(p, row_t)

	out_path = Path(file_path).parent / "translated.xlsx"
	trans.save(str(out_path))


if __name__ == "__main__":
	asyncio.run(main("/app/./scripts/taxonomy-with-ids.ru-RU.xlsx"))
